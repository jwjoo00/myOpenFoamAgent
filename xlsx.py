"""
xlsx.py — build an .xlsx workbook with tables + embedded images (openpyxl).

Backs the build_excel tool. Each sheet spec is:
    {"name": "...",
     "tables": [{"title": "...", "columns": [...], "rows": [[...], ...]}],
     "images": ["/path/to/fig.png", ...]}

No LLM dependency. Image embedding needs Pillow (pulled in by matplotlib).
"""
from __future__ import annotations

from pathlib import Path


def _cell(v):
    if v is None:
        return ""
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float, str)):
        return v
    return str(v)


def build(dest, sheets, confine=None) -> dict:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    dest = Path(dest)
    if not isinstance(sheets, list) or not sheets:
        sheets = [{"name": "Sheet1"}]

    wb = Workbook()
    default = wb.active
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(bold=True, color="FFFFFF")
    n_tables = n_images = 0

    for si, sh in enumerate(sheets):
        sh = sh if isinstance(sh, dict) else {}
        name = (str(sh.get("name") or f"Sheet{si + 1}"))[:31]
        ws = wb.create_sheet(title=name)
        r = 1
        max_col = 1

        for tbl in (sh.get("tables") or []):
            if not isinstance(tbl, dict):
                continue
            title = tbl.get("title")
            cols = [str(c) for c in (tbl.get("columns") or [])]
            rows = tbl.get("rows") or []
            if title:
                ws.cell(row=r, column=1, value=str(title)).font = Font(bold=True, size=12)
                r += 1
            if cols:
                for c, nm in enumerate(cols, 1):
                    cell = ws.cell(row=r, column=c, value=nm)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                max_col = max(max_col, len(cols))
                r += 1
            for row in rows:
                row = row if isinstance(row, (list, tuple)) else [row]
                for c, v in enumerate(row, 1):
                    ws.cell(row=r, column=c, value=_cell(v))
                max_col = max(max_col, len(row))
                r += 1
            n_tables += 1
            r += 1   # spacer between tables

        for c in range(1, max_col + 1):     # column widths from content
            w = 10
            for rr in range(1, r):
                v = ws.cell(row=rr, column=c).value
                if v is not None:
                    w = max(w, len(str(v)))
            ws.column_dimensions[get_column_letter(c)].width = min(w + 2, 60)

        for img in (sh.get("images") or []):    # images stacked below the tables
            ip = Path(str(img))
            if confine is not None:
                cp = confine(img)
                if cp is None:
                    continue
                ip = cp
            if not ip.is_file():
                continue
            try:
                xi = XLImage(str(ip))
                ws.add_image(xi, f"A{r}")
                r += max(20, int((getattr(xi, "height", 360) or 360) / 18) + 2)
                n_images += 1
            except Exception:
                continue

    if default in wb.worksheets and len(wb.worksheets) > 1:
        wb.remove(default)      # drop the auto "Sheet" once named sheets exist

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(dest))
    return {"ok": True, "path": str(dest), "sheets": wb.sheetnames,
            "tables": n_tables, "images": n_images}
