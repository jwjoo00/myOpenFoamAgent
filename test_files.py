#!/usr/bin/env python3
"""test_files.py — write_file path confinement + build_excel (tables+image). NO LLM."""
from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import config


def main() -> int:
    with tempfile.TemporaryDirectory() as d:
        config.RUNS_ROOT = Path(d) / "runs"
        config.RUNS_ROOT.mkdir()
        config.REPORTS_ROOT = Path(d) / "reports"
        config.REPORTS_ROOT.mkdir()

        import tools
        ok = True

        # write_file inside an allowed root
        r = tools.write_file(str(config.RUNS_ROOT / "case1" / "note.txt"), "hello\n표 데이터")
        ok &= r["ok"] and Path(r["path"]).read_text() == "hello\n표 데이터"

        # escapes rejected
        bad1 = tools.write_file("/etc/oops.txt", "x")
        bad2 = tools.write_file(str(config.RUNS_ROOT / ".." / "escape.txt"), "x")
        ok &= (not bad1["ok"]) and (not bad2["ok"])
        print(f"[write_file] write_ok={r['ok']} escapes_rejected={not bad1['ok'] and not bad2['ok']}")

        # make a png inside reports/ to embed
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        png = config.REPORTS_ROOT / "chart.png"
        plt.figure(); plt.plot([1, 2, 3], [1, 4, 9]); plt.savefig(png); plt.close()

        # build_excel: a comparison table + the image
        xp = config.REPORTS_ROOT / "compare.xlsx"
        rr = tools.build_excel(str(xp), [{
            "name": "Re sweep",
            "tables": [{"title": "Reynolds comparison",
                        "columns": ["Re", "nu", "Cd", "cells"],
                        "rows": [[100, "1e-3", 1.21, 400], [400, "2.5e-4", 1.10, 1600]]}],
            "images": [str(png)],
        }])
        ok &= rr.get("ok") and Path(rr["path"]).is_file()
        print(f"[build_excel] {rr}")

        # image outside roots must be skipped (not embedded), not crash
        rr2 = tools.build_excel(str(config.REPORTS_ROOT / "x2.xlsx"),
                                [{"name": "S", "images": ["/etc/hosts"]}])
        ok &= rr2.get("ok") and rr2.get("images") == 0

        # read back the workbook
        from openpyxl import load_workbook
        wb = load_workbook(str(xp))
        ws = wb["Re sweep"]
        flat = [v for row in ws.iter_rows(values_only=True) for v in row]
        ok &= "Reynolds comparison" in flat and 100 in flat and 1.21 in flat
        ok &= len(getattr(ws, "_images", [])) == 1
        print(f"[xlsx readback] title={'Reynolds comparison' in flat} "
              f"re100={100 in flat} images={len(ws._images)}")

        print("\nRESULT:", "PASS" if ok else "FAIL")
        return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
